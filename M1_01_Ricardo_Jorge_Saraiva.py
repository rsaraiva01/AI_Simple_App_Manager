import openpyxl
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("stock_fruta.xlsx")
sheetname = wb.sheetnames[0] #variável para a 1.ª folha
sheet_1 = wb[sheetname]

from fruit_synonyms import fruit_syntax, fruit_categories 

print(f"\t\tType 'exit' to close the chat.")
print(f"\t\tType 'help' for an explanation of the program.")
print(f"\n\t\tHello! Which fruit would you like to check in stock?\n")

should_exit = False

class Chatbot():
    exit_words = ["exit", "Exit"]
    should_exit = False

    def __init__(self):
        pass
        #self.user_message = user_message          

    def read_and_format(self):
        user_message = input("You: ")
        if user_message in self.exit_words:
            self.should_exit = True
            return self.should_exit
        formated_message = user_message.lower().split()
        #Check EMPTY message
        if not user_message:
            return
        return formated_message

    #Search in Fruit Syntax Dict
    def search_fruit_synonyms(self, formatted_message):
        fruits_found = []
        for user_word in formatted_message:  # user message
            for fruit, synonyms in fruit_syntax.items():  # dict
                if user_word == fruit or user_word in synonyms:
                    fruits_found.append(fruit)
        #print("\t", fruits_found)
        return fruits_found

    def search_fruit_db(self, fruits_found):
        searched_fruits_stock = {}
        for fruit_name in fruits_found:
            #print(f"\tVerifying stock for {fruit_name}...")
            line = 0
            max_row = sheet_1.max_row
            while line < max_row:
                search_fruits_result = sheet_1.cell(row = line + 1, column = 1).value        
                if fruit_name == search_fruits_result:  
                    search_stock_units =  sheet_1.cell(row = line + 1, column = 2).value            
                    searched_fruits_stock[search_fruits_result] = search_stock_units
                    line = max_row #Found it! Skip to search another word
                line = line + 1
        if searched_fruits_stock:
            print("\t\tFruit(s) stock:", searched_fruits_stock) 
        return searched_fruits_stock

    #Search in Fruit Categories Dict
    def suggest_alternative(self, searched_fruits_stock):
        alternatives = {}

        # Check for ZERO units
        for fruit_name, stock_unit in searched_fruits_stock.items():
            if stock_unit == 0:
                print(f"\t\tOut of stock for {fruit_name}.")
                
                # Find Category
                for category_name, category_fruit_name in fruit_categories.items():
                    if fruit_name in category_fruit_name:
                        category = category_name
                        break
                
                # Find fruits in Category
                if category:
                    print(f"\t\tFinding alternatives related to {category}...")

                    # List of fruits in category
                    potential_alternatives = fruit_categories[category]
                    # Check fruit by fruit in category
                    for fruit_alternative in potential_alternatives:

                        # Exclude customer searched fruit before searching in Excel
                        if fruit_alternative in searched_fruits_stock:
                            continue

                        alternative_stock = 0
                        line = 0
                        max_row = sheet_1.max_row

                        while line < max_row:
                            fruit_name = sheet_1.cell(row = line + 1, column = 1).value 
                            
                            if fruit_name and fruit_alternative == fruit_name.lower():
                                stock_units = sheet_1.cell(row = line + 1, column = 2).value
                                # Certificamo-nos que o stock é 0 se estiver vazio
                                alternative_stock = stock_units if stock_units is not None else 0
                                break # Stock found!
                            
                            line += 1
                        
                        # Suggest only if stock units > 0
                        if alternative_stock > 0:
                            alternatives[fruit_alternative] = alternative_stock

        # Print alternatives
        if alternatives:
            print(f"\n\t\tAlternative(s): {alternatives}")
        elif any(stock == 0 for stock in searched_fruits_stock.values()):
            print(f"\t\tAlternatives not found in same category.")
            
        return alternatives


#-----------------------------------------------------
while should_exit == False:
    customer_message = Chatbot()
    message = customer_message.read_and_format()
    print("")
    ai_syntax = customer_message.search_fruit_synonyms(message)
    search_stock = customer_message.search_fruit_db(ai_syntax)
    if ai_syntax:
        alternatives = customer_message.suggest_alternative(search_stock)
    else:
        print(f"\t\tHow can i help you today?")

    print("")


    if customer_message.should_exit:
        print(f"\t\tThank you.")
        should_exit = True
        break
